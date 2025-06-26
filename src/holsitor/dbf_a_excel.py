import shutil
import os
import re
from dbfread import DBF
from openpyxl import Workbook
from datetime import datetime

def limpiar_nombre_columna(nombre, index=None, errores_log=[]):
    """
    Reemplaza caracteres inválidos en nombres de columnas y asigna un nombre seguro si está vacío o ilegible.
    Registra cualquier cambio en el log de errores.
    """
    original = str(nombre)
    limpio = re.sub(r'[\\/*?:[\]☻]', '_', original.strip())

    # Si está vacío o contiene solo símbolos inválidos
    if not limpio or limpio in ['_', '☻']:
        limpio = f"COL_{index}"
        errores_log.append(f"⚠️ Columna inválida en posición {index}: '{original}' → renombrada como '{limpio}'")
    elif original != limpio:
        errores_log.append(f"⚠️ Columna renombrada: '{original}' → '{limpio}'")

    return limpio

def hoja_segura(wb: Workbook, tipo: str):
    """
    Asigna un nombre seguro a la hoja activa del workbook.
    Si hay error, asigna 'Datos'.
    """
    ws = wb.active
    try:
        ws.title = tipo if tipo in ["Ventas", "Compras"] else "Datos"
    except:
        ws.title = "Datos"
    return ws

def convertir_dbf_a_excel(dbf_filename: str, tipo: str, via: str, output_folder: str, client_name: str) -> str:
    """
    Copia un archivo DBF desde la vía Holistor, lo convierte a Excel y lo guarda en la carpeta del cliente.

    Args:
        dbf_filename (str): "MOVIVAV.DBF" o "MOVIVAC.DBF"
        tipo (str): "Ventas" o "Compras"
        via (str): nombre de la carpeta Holistor (ej: "PEREZSOC")
        output_folder (str): carpeta de emisión
        client_name (str): nombre del cliente

    Returns:
        str: ruta al archivo Excel generado o None si falló
    """
    errores_columnas = []

    try:
        # Ruta al archivo DBF original
        base_path = r"C:\HolistorW\Whdatos"
        original_file = os.path.join(base_path, via, dbf_filename)

        if not os.path.exists(original_file):
            raise FileNotFoundError(f"No se encontró el archivo DBF: {original_file}")

        # Copiar archivo DBF a la carpeta de emisión
        copia_file = os.path.join(output_folder, f"COPIA_{tipo.upper()}_{client_name}.DBF")
        shutil.copy(original_file, copia_file)

        # Leer el DBF
        dbf_table = DBF(copia_file, load=True, encoding='latin-1')

        # Crear workbook sin hoja activa
        wb = Workbook()
        wb.remove(wb.active)

        # Forzar nombre seguro
        nombre_hoja = "Compras" if "COMPR" in dbf_filename.upper() else "Ventas"
        try:
            ws = wb.create_sheet(title=nombre_hoja)
        except:
            ws = wb.create_sheet(title="Datos")


        # Limpiar encabezados
        clean_headers = [limpiar_nombre_columna(col, i, errores_columnas) for i, col in enumerate(dbf_table.field_names)]
        ws.append(clean_headers)

        # Escribir filas
        for record in dbf_table:
            ws.append([record.get(field, "") for field in dbf_table.field_names])

        # Guardar Excel
        excel_filename = f"TEMP - {tipo} - {client_name}.xlsx"
        excel_path = os.path.join(output_folder, excel_filename)
        wb.save(excel_path)

        print(f"✅ {tipo} convertido a Excel: {excel_path}")

        if errores_columnas:
            print(f"📝 Log de columnas corregidas en {tipo}:")
            for error in errores_columnas:
                print("   " + error)

        return excel_path

    except Exception as e:
        print(f"❌ Error al convertir {tipo} para {client_name}: {e}")
        return None
